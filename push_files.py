from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Iterable

import oracledb
import pandas as pd

from process_files import ProcessedFile, find_files, get_file_details, process_files


ORACLE_IDENTIFIER = re.compile(r"^[A-Za-z][A-Za-z0-9_$#]*$")
DATE_ADDED_COLUMN = "DATE_ADDED"
UPLOAD_LOG_FILE = "upload_log.txt"
DAILY_STATUS_FILE = "daily_file_status.xlsx"
ACTION_CREATED_TABLE = "Created table"
ACTION_APPENDED = "Appended"
ACTION_SKIPPED = "Skipped"

# Add Oracle reserved or project-specific column replacements here.
COLUMN_NAME_REPLACEMENTS = {
    "DATE": "EVENT_DATE",
}


def push_files(
    file_paths: str | Path | Iterable[str | Path],
    oracle_user: str,
    oracle_password: str,
    oracle_dsn: str,
    uploaded_directory: str | Path | None = None,
    log_directory: str | Path | None = None,
) -> None:
    """Insert files into Oracle, then optionally move them to an archive folder."""
    source_files = _get_file_paths(file_paths)
    attempted_files = list(source_files)
    skipped_log_entries = []

    if uploaded_directory:
        source_files, skipped_files = _split_archived_files(
            source_files,
            uploaded_directory,
        )
        archive_files(skipped_files, uploaded_directory, force_rename=True)
        skipped_log_entries.extend(_get_skipped_log_entries(skipped_files))

    if log_directory:
        source_files, skipped_files = _split_logged_files(source_files, log_directory)

        if uploaded_directory:
            archive_files(skipped_files, uploaded_directory, force_rename=True)

        skipped_log_entries.extend(_get_skipped_log_entries(skipped_files))
        append_upload_log(skipped_log_entries, log_directory)

    processed_files = process_files(source_files)
    if not processed_files:
        if log_directory:
            safe_update_daily_status_log(attempted_files, [], log_directory)
        return

    with oracledb.connect(
        user=oracle_user,
        password=oracle_password,
        dsn=oracle_dsn,
    ) as connection:
        upload_log_entries = push_processed_files(connection, processed_files)

    if log_directory:
        append_upload_log(upload_log_entries, log_directory)
        safe_update_daily_status_log(attempted_files, upload_log_entries, log_directory)

    if uploaded_directory:
        archive_processed_files(processed_files, uploaded_directory)


def _get_file_paths(
    file_paths: str | Path | Iterable[str | Path],
) -> list[Path]:
    """Return files from a directory path, a single file path, or many paths."""
    if isinstance(file_paths, (str, Path)):
        path = Path(file_paths)

        if path.is_dir():
            return find_files(path)

        return [path]

    return [Path(path) for path in file_paths]


def _split_archived_files(
    file_paths: list[Path],
    uploaded_directory: str | Path,
) -> tuple[list[Path], list[Path]]:
    """Separate new files from files whose names already exist in the archive."""
    uploaded_path = Path(uploaded_directory)
    uploaded_path.mkdir(parents=True, exist_ok=True)
    new_files = []
    skipped_files = []

    for file_path in file_paths:
        if _is_inside_directory(file_path, uploaded_path):
            continue

        if (uploaded_path / file_path.name).exists():
            print(f"Skipped {file_path.name}: already uploaded to Oracle")
            skipped_files.append(file_path)
        else:
            new_files.append(file_path)

    return new_files, skipped_files


def _split_logged_files(
    file_paths: list[Path],
    log_directory: str | Path,
) -> tuple[list[Path], list[Path]]:
    """Separate new files from files already recorded in the upload log."""
    uploaded_file_names = read_logged_file_names(log_directory)
    new_files = []
    skipped_files = []

    for file_path in file_paths:
        if file_path.name in uploaded_file_names:
            print(f"Skipped {file_path.name}: already found in upload log")
            skipped_files.append(file_path)
        else:
            new_files.append(file_path)

    return new_files, skipped_files


def read_logged_file_names(log_directory: str | Path) -> set[str]:
    """Return file names already present in the upload log."""
    log_path = _get_upload_log_path(log_directory)
    if not log_path.exists():
        return set()

    uploaded_file_names = set()

    for line in log_path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue

        file_name = _get_logged_file_name(line)
        if file_name:
            uploaded_file_names.add(file_name)

    return uploaded_file_names


def append_upload_log(
    upload_log_entries: list[tuple[str, str, str | None]],
    log_directory: str | Path,
) -> None:
    """Append upload or skip entries to the upload log file."""
    if not upload_log_entries:
        return

    log_path = _get_upload_log_path(log_directory)
    log_path.parent.mkdir(parents=True, exist_ok=True)

    with log_path.open("a", encoding="utf-8") as log_file:
        for file_name, action, table_name in upload_log_entries:
            log_file.write(_format_upload_log_entry(file_name, action, table_name))


def update_daily_status_log(
    attempted_files: Iterable[str | Path],
    upload_log_entries: list[tuple[str, str, str | None]],
    log_directory: str | Path,
) -> None:
    """Update the Y/N workbook using dates parsed from file names."""
    status_path = Path(log_directory) / DAILY_STATUS_FILE
    status_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = _open_status_workbook(status_path)
    worksheet = workbook.active
    worksheet.title = "Daily Status"

    _ensure_status_headers(worksheet)
    attempted_status = _get_attempted_status_entries(attempted_files)
    successful_status = _get_successful_status_entries(upload_log_entries)

    for file_group, file_date in sorted(attempted_status | successful_status):
        date_column = _get_or_create_date_column(worksheet, file_date)
        row_number = _get_or_create_status_row(worksheet, file_group)

        if worksheet.cell(row=row_number, column=date_column).value is None:
            worksheet.cell(row=row_number, column=date_column).value = "N"

        if (file_group, file_date) in successful_status:
            worksheet.cell(row=row_number, column=date_column).value = "Y"

    workbook.save(status_path)


def safe_update_daily_status_log(
    attempted_files: Iterable[str | Path],
    upload_log_entries: list[tuple[str, str, str | None]],
    log_directory: str | Path,
) -> None:
    """Update the daily workbook without blocking archive cleanup."""
    try:
        update_daily_status_log(attempted_files, upload_log_entries, log_directory)
    except Exception as error:
        print(f"Could not update daily status workbook: {error}")


def _open_status_workbook(status_path: Path):
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Daily status workbook requires openpyxl. "
            "Install dependencies with: pip install -r requirements.txt"
        ) from error

    if status_path.exists():
        return load_workbook(status_path)

    return Workbook()


def _ensure_status_headers(worksheet) -> None:
    if worksheet.cell(row=1, column=1).value != "Files":
        worksheet.cell(row=1, column=1).value = "Files"


def _get_or_create_date_column(worksheet, file_date: str) -> int:
    for column_number in range(2, worksheet.max_column + 1):
        if worksheet.cell(row=1, column=column_number).value == file_date:
            return column_number

    column_number = worksheet.max_column + 1
    worksheet.cell(row=1, column=column_number).value = file_date

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_number, column=column_number).value = "N"

    return column_number


def _get_attempted_status_entries(
    attempted_files: Iterable[str | Path],
) -> set[tuple[str, str]]:
    return {
        _get_file_status_key(file_path)
        for file_path in attempted_files
    }


def _get_successful_status_entries(
    upload_log_entries: list[tuple[str, str, str | None]],
) -> set[tuple[str, str]]:
    return {
        _get_file_status_key(file_name)
        for file_name, action, _table_name in upload_log_entries
        if action != ACTION_SKIPPED
    }


def _get_or_create_status_row(worksheet, file_group: str) -> int:
    for row_number in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row_number, column=1).value == file_group:
            return row_number

    row_number = worksheet.max_row + 1
    worksheet.cell(row=row_number, column=1).value = file_group

    for column_number in range(2, worksheet.max_column + 1):
        worksheet.cell(row=row_number, column=column_number).value = "N"

    return row_number


def _get_file_status_key(file_path: str | Path) -> tuple[str, str]:
    table_name, file_date = get_file_details(file_path)
    return table_name, _format_file_date(file_date)


def _format_file_date(file_date: str) -> str:
    return datetime.strptime(file_date, "%Y%m%d").strftime("%Y-%m-%d")


def _get_skipped_log_entries(
    file_paths: Iterable[Path],
) -> list[tuple[str, str, str | None]]:
    return [(file_path.name, ACTION_SKIPPED, None) for file_path in file_paths]


def _get_logged_file_name(line: str) -> str:
    log_entry = line.strip()
    if " : " in log_entry:
        log_entry = log_entry.split(" : ", maxsplit=1)[1].strip()

    return log_entry.split(" -> ", maxsplit=1)[0].strip()


def _format_upload_log_entry(
    file_name: str,
    action: str,
    table_name: str | None,
) -> str:
    logged_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"{logged_at} : {file_name} -> {action}"

    if table_name:
        log_entry = f"{log_entry} -> {table_name}"

    return f"{log_entry}\n"


def _get_upload_log_path(log_directory: str | Path) -> Path:
    return Path(log_directory) / UPLOAD_LOG_FILE


def _is_inside_directory(file_path: Path, directory: Path) -> bool:
    """Return whether one file is already inside a directory."""
    try:
        file_path.resolve().relative_to(directory.resolve())
        return True
    except ValueError:
        return False


def push_processed_files(
    connection,
    processed_files: dict[str, list[ProcessedFile]],
) -> list[tuple[str, str, str | None]]:
    """Insert all processed files and commit once everything succeeds."""
    upload_log_entries = []

    try:
        for table_name, files_for_table in processed_files.items():
            for processed_file in files_for_table:
                action = push_dataframe(connection, table_name, processed_file.data)
                if action:
                    upload_log_entries.append(
                        (processed_file.file_path.name, action, table_name)
                    )
                print(
                    f"Inserted {len(processed_file.data)} rows from "
                    f"{processed_file.file_path.name} into {table_name}"
                )

        connection.commit()
        return upload_log_entries
    except Exception:
        connection.rollback()
        raise


def archive_processed_files(
    processed_files: dict[str, list[ProcessedFile]],
    uploaded_directory: str | Path,
) -> None:
    """Move successfully uploaded source files into the archive folder."""
    file_paths = [
        processed_file.file_path
        for files_for_table in processed_files.values()
        for processed_file in files_for_table
    ]
    archive_files(file_paths, uploaded_directory)


def archive_files(
    file_paths: Iterable[str | Path],
    uploaded_directory: str | Path,
    force_rename: bool = False,
) -> None:
    """Move source files into the archive folder without overwriting older files."""
    uploaded_path = Path(uploaded_directory)
    uploaded_path.mkdir(parents=True, exist_ok=True)

    for file_path in file_paths:
        source_path = Path(file_path)
        destination = _get_archive_destination(
            uploaded_path,
            source_path.name,
            force_rename=force_rename,
        )
        shutil.move(str(source_path), str(destination))
        print(f"Moved {source_path.name} to {destination}")


def _get_archive_destination(
    uploaded_directory: Path,
    file_name: str,
    force_rename: bool = False,
) -> Path:
    """Return an unused archive path without overwriting an older upload."""
    destination = uploaded_directory / file_name
    if not force_rename and not destination.exists():
        return destination

    file_path = Path(file_name)
    uploaded_at = datetime.now().strftime("%Y%m%d_%H%M%S")
    destination = uploaded_directory / f"{file_path.stem}_{uploaded_at}{file_path.suffix}"
    duplicate_number = 1

    while destination.exists():
        destination = uploaded_directory / (
            f"{file_path.stem}_{uploaded_at}_{duplicate_number}{file_path.suffix}"
        )
        duplicate_number += 1

    return destination


def push_dataframe(connection, table_name: str, dataframe: pd.DataFrame) -> str | None:
    """Create the target table when needed, then insert one DataFrame."""
    if dataframe.empty:
        return None

    _validate_identifier(table_name)
    columns = _sanitize_column_names(dataframe.columns)
    _validate_date_added_column(columns)

    if not table_exists(connection, table_name):
        create_table(connection, table_name, dataframe)
        action = ACTION_CREATED_TABLE
    elif not table_column_exists(connection, table_name, DATE_ADDED_COLUMN):
        add_date_added_column(connection, table_name)
        action = ACTION_APPENDED
    else:
        action = ACTION_APPENDED

    insert_columns = [*columns, DATE_ADDED_COLUMN]
    column_list = ", ".join(insert_columns)
    bind_variables = ", ".join(
        f":{number}" for number in range(1, len(insert_columns) + 1)
    )
    insert_sql = f"INSERT INTO {table_name} ({column_list}) VALUES ({bind_variables})"

    date_added = datetime.now()
    rows = [
        tuple(_to_oracle_value(value) for value in row) + (date_added,)
        for row in dataframe.itertuples(index=False, name=None)
    ]

    with connection.cursor() as cursor:
        cursor.executemany(insert_sql, rows)

    return action


def table_exists(connection, table_name: str) -> bool:
    """Return whether the current Oracle user already owns the table."""
    _validate_identifier(table_name)

    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT COUNT(*) FROM user_tables WHERE table_name = :1",
            [table_name.upper()],
        )
        return cursor.fetchone()[0] > 0


def table_column_exists(connection, table_name: str, column_name: str) -> bool:
    """Return whether an Oracle table already contains one column."""
    _validate_identifier(table_name)
    _validate_identifier(column_name)

    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT COUNT(*)
            FROM user_tab_columns
            WHERE table_name = :1 AND column_name = :2
            """,
            [table_name.upper(), column_name.upper()],
        )
        return cursor.fetchone()[0] > 0


def add_date_added_column(connection, table_name: str) -> None:
    """Add the load timestamp column to an existing Oracle table."""
    _validate_identifier(table_name)

    with connection.cursor() as cursor:
        cursor.execute(f"ALTER TABLE {table_name} ADD {DATE_ADDED_COLUMN} DATE")

    print(f"Added {DATE_ADDED_COLUMN} column to {table_name}")


def create_table(connection, table_name: str, dataframe: pd.DataFrame) -> None:
    """Create an Oracle table using the columns and types in one DataFrame."""
    _validate_identifier(table_name)
    column_definitions = []
    column_names = _sanitize_column_names(dataframe.columns)
    _validate_date_added_column(column_names)

    for column, column_name in zip(dataframe.columns, column_names):
        oracle_type = _get_oracle_type(dataframe[column])
        column_definitions.append(f"{column_name} {oracle_type}")

    column_definitions.append(f"{DATE_ADDED_COLUMN} DATE")
    create_sql = f"CREATE TABLE {table_name} ({', '.join(column_definitions)})"

    with connection.cursor() as cursor:
        cursor.execute(create_sql)

    print(f"Created table {table_name}")


def _get_oracle_type(series: pd.Series) -> str:
    """Map a pandas column to a simple Oracle data type."""
    if pd.api.types.is_bool_dtype(series):
        return "CHAR(1)"
    """
    # Commented these so these should go as VARCHAR2, but can be easily changed back if needed.
    if pd.api.types.is_integer_dtype(series):
        return "NUMBER"

    if pd.api.types.is_float_dtype(series):
        return "NUMBER"
    """
    if pd.api.types.is_datetime64_any_dtype(series):
        return "DATE"

    max_length = series.dropna().astype(str).str.len().max()
    max_length = 1 if pd.isna(max_length) else int(max_length)

    if max_length > 4000:
        return "CLOB"

    max_length = max(255, max_length)
    return f"VARCHAR2({max_length})"


def _sanitize_column_names(columns) -> list[str]:
    """Return unique Oracle-friendly column names."""
    sanitized_columns = [_sanitize_column_name(column) for column in columns]

    if len(sanitized_columns) != len(set(sanitized_columns)):
        raise ValueError(
            "Two or more source columns become the same Oracle column name after "
            f"sanitization: {sanitized_columns}"
        )

    return sanitized_columns


def _validate_date_added_column(columns: list[str]) -> None:
    if DATE_ADDED_COLUMN in columns:
        raise ValueError(
            f"Source files cannot contain the reserved column: {DATE_ADDED_COLUMN}"
        )


def _to_oracle_value(value):
    """Convert pandas values into values accepted by Oracle."""
    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    return value


def _sanitize_column_name(column) -> str:
    """Convert one source header into an Oracle-friendly column name."""
    column_name = re.sub(r"[^A-Za-z0-9_$#]+", "_", str(column))
    column_name = column_name.strip("_").upper()

    if not column_name:
        raise ValueError(f"Could not create an Oracle column name from: {column}")

    if not column_name[0].isalpha():
        column_name = f"COLUMN_{column_name}"

    column_name = COLUMN_NAME_REPLACEMENTS.get(column_name, column_name)

    _validate_identifier(column_name)
    return column_name


def _validate_identifier(identifier: str) -> None:
    if not ORACLE_IDENTIFIER.fullmatch(identifier):
        raise ValueError(f"Invalid Oracle identifier: {identifier}")
